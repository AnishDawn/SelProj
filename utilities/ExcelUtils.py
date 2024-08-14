import os
from typing import List, Any

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException


def get_file_path() -> str:
    base_dir = os.path.dirname(os.path.abspath(__file__))  # Directory of the current script
    return os.path.join(base_dir, '..', 'ExcelFiles', 'testData.xlsx')


def load_workbook(path: str, sheet_name: str):
    """Loads the workbook and returns the specified sheet."""
    try:
        workbook = openpyxl.load_workbook(path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")
        sheet = workbook[sheet_name]
        return workbook, sheet
    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{path}' was not found.")
    except InvalidFileException:
        raise InvalidFileException(f"The file '{path}' is not a valid Excel file.")
    except Exception as e:
        raise Exception(f"An error occurred while loading the workbook: {e}")


def get_row_count(path: str, sheet_name: str) -> int:
    """Returns the number of rows in the specified sheet."""
    _, sheet = load_workbook(path, sheet_name)  # Load the workbook and get the sheet, ignoring the workbook
    return sheet.max_row


def get_column_count(path: str, sheet_name: str) -> int:
    """Returns the number of columns in the specified sheet."""
    _, sheet = load_workbook(path, sheet_name)
    return sheet.max_column


def get_cell_data(path: str, sheet_name: str, row_number: int, column_number: int) -> Any:
    """Returns the data from a specified cell."""
    try:
        _, sheet = load_workbook(path, sheet_name)
        return sheet.cell(row=row_number, column=column_number).value
    except IndexError:
        raise IndexError(f"Cell at row {row_number} and column {column_number} is out of range.")
    except Exception as e:
        raise Exception(f"An error occurred while getting cell data: {e}")


def set_cell_data(path: str, sheet_name: str, row_number: int, column_number: int, data: Any) -> None:
    """Sets the data in a specified cell and saves the workbook."""
    try:
        workbook, sheet = load_workbook(path, sheet_name)
        sheet.cell(row=row_number, column=column_number).value = data
        with open(path, "wb") as f:
            workbook.save(f)
    except IndexError:
        raise IndexError(f"Cell at row {row_number} and column {column_number} is out of range.")
    except Exception as e:
        raise Exception(f"An error occurred while setting cell data: {e}")


def get_data_from_excel(path: str, sheet_name: str) -> List[List[Any]]:
    """Returns all data from the specified sheet as a list of lists."""
    try:
        _, sheet = load_workbook(path, sheet_name)
        total_rows = sheet.max_row
        total_cols = sheet.max_column

        final_list = []
        for r in range(2, total_rows + 1):
            row_list = [sheet.cell(row=r, column=c).value for c in range(1, total_cols + 1)]
            final_list.append(row_list)

        return final_list
    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{path}' was not found.")
    except InvalidFileException:
        raise InvalidFileException(f"The file '{path}' is not a valid Excel file.")
    except Exception as e:
        raise Exception(f"An error occurred while getting data from the Excel sheet: {e}")
